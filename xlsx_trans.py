# -*- coding: utf-8 -*-
# !/usr/bin/env python
# --author-- peibibing
from openpyxl import load_workbook
import datetime, csv, re
# from multiprocessing import Pool
from multiprocessing.dummy import Pool

# def complete_data():










#包装
def xlsx_trans(wb):
    # wb = load_workbook('/Users/wode/Downloads/test.xlsx')
    # name = wb.sheetnames[0]
    # ws = wb.active
    ws = wb.get_sheet_by_name(wb.sheetnames[4])
    n1 = ws.title
    # n2 = re.search('\d{1,6}', j).group()
    n2 = re.findall('^\d*', n1)[0]
    h = 6 - len(n2) - 1
    h1 = ''
    for i in range(h):
        h1 += '0'
    name = 'H' + h1 + n2
    # print (name)
    # return 0

    row_num = ws.max_row
    # col_num = ws.max_column

    # 补全数据
    data_start = ws.rows[4][0].value  # 表中开始时间
    data_end = ws.rows[row_num - 1][0].value  # 表中最后结束时间
    period_time = (data_end - data_start).seconds / 60 + 1  # 中间时间间隔
    # open1 = []
    period_time1 = range(period_time)
    i = 0
    for row in range(row_num):
        if i > period_time:
            break
        if row >= row_num - 4:
            break
        date = ws.rows[row + 4][0].value
        data_start1 = data_start + datetime.timedelta(minutes=period_time1[i])
        i += 1
        if date == data_start1:
            if date < datetime.datetime(2015, 12, 15, 9, 30):
                continue
            else:
                date1 = ws.rows[row + 4][0].value.strftime("%Y%m%d%H%M")
                open1 = ws.rows[row + 4][1].value
                high = ws.rows[row + 4][2].value
                low = ws.rows[row + 4][3].value
                last_price = ws.rows[row + 4][4].value
                number_ticks = ws.rows[row + 4][5].value
                volume = ws.rows[row + 4][6].value
                value = ws.rows[row + 4][7].value
                yd = date1[:8]
                time = date1[8:]
                H = name + date1
                with open('/Users/wode/Desktop/test.csv', 'a')as f:
                    writer = csv.writer(f)
                    writer.writerow([H, name, yd, time, open1, high, low, last_price, number_ticks, volume, value])
                print(i)
        else:
            while date > data_start1:

                date1 = data_start1.strftime("%Y%m%d%H%M")
                open1 = ws.rows[row + 3][1].value
                high = ws.rows[row + 3][2].value
                low = ws.rows[row + 3][3].value
                last_price = ws.rows[row + 3][4].value
                yd = date1[:8]
                time = date1[8:]
                H = name + date1
                with open('/Users/wode/Desktop/test.csv', 'a')as f:
                    writer = csv.writer(f)
                    writer.writerow([H, name, yd, time, open1, high, low, last_price, 0, 0, 0])
                data_start1 = data_start1 + datetime.timedelta(minutes=1)
                i += 1
                print(i)
            else:
                date1 = data_start1.strftime("%Y%m%d%H%M")
                open1 = ws.rows[row + 4][1].value
                high = ws.rows[row + 4][2].value
                low = ws.rows[row + 4][3].value
                last_price = ws.rows[row + 4][4].value
                number_ticks = ws.rows[row + 4][5].value
                volume = ws.rows[row + 4][6].value
                value = ws.rows[row + 4][7].value
                yd = date1[:8]
                time = date1[8:]
                H = name + date1
                with open('/Users/wode/Desktop/test.csv', 'a')as f:
                    writer = csv.writer(f)
                    writer.writerow([H, name, yd, time, open1, high, low, last_price, number_ticks, volume, value])
                print(i)
    return 0
if __name__ == '__main__':
    wb1 = load_workbook('/Users/wode/Downloads/航空 (1).xlsx')
    pool = Pool(20)
    pool.map(xlsx_trans, [wb1])
    # a = xlsx_trans(wb1)



# t1 = datetime.datetime.now()
# wb1 = load_workbook('/Users/wode/Downloads/航空 (1).xlsx')
# t2 = datetime.datetime.now()
# print(t2-t1)



# wb = load_workbook('/Users/wode/Downloads/test.xlsx')
# # name = wb.sheetnames[0]
# ws = wb.active
# n1 = ws.title
# #n2 = re.search('\d{1,6}', j).group()
# n2 = re.findall('^\d*', n1)[0]
# h = 6 - len(n2) - 1
# h1 = ''
# for i in range(h):
#     h1 += '0'
# name = 'H' + h1 + n2
#
# row_num = ws.max_row
# col_num = ws.max_column
#
# #补全数据
# data_start = ws.rows[4][0].value#表中开始时间
# data_end = ws.rows[row_num-1][0].value#表中最后结束时间
# period_time = (data_end-data_start).seconds/60 + 1#中间时间间隔
# open1 = []
# period_time1 = range(period_time)
# i = 0
# for row in range(row_num):
#     if i > period_time:
#         break
#     if row >= row_num - 4:
#         break
#     date = ws.rows[row+4][0].value
#     data_start1 = data_start + datetime.timedelta(minutes=period_time1[i])
#     i += 1
#     if date == data_start1:
#         if date < datetime.datetime(2015, 12, 15, 9, 30):
#             continue
#         else:
#             date1 = ws.rows[row+4][0].value.strftime("%Y%m%d%H%M")
#             open1 = ws.rows[row+4][1].value
#             high = ws.rows[row+4][2].value
#             low = ws.rows[row+4][3].value
#             last_price = ws.rows[row+4][4].value
#             number_ticks = ws.rows[row+4][5].value
#             volume = ws.rows[row+4][6].value
#             value = ws.rows[row+4][7].value
#             yd = date1[:8]
#             time = date1[8:]
#             H = name+date1
#             with open('/Users/wode/Desktop/test.csv', 'a')as f:
#                 writer = csv.writer(f)
#                 writer.writerow([H, name, yd, time, open1, high, low, last_price, number_ticks, volume, value])
#     else:
#         while date > data_start1:
#
#             date1 = data_start1.strftime("%Y%m%d%H%M")
#             open1 = ws.rows[row+3][1].value
#             high = ws.rows[row+3][2].value
#             low = ws.rows[row+3][3].value
#             last_price = ws.rows[row+3][4].value
#             yd = date1[:8]
#             time = date1[8:]
#             H = name + date1
#             with open('/Users/wode/Desktop/test.csv', 'a')as f:
#                 writer = csv.writer(f)
#                 writer.writerow([H, name, yd, time, open1, high, low, last_price, 0, 0, 0])
#             data_start1 = data_start1 + datetime.timedelta(minutes=1)
#             i += 1
#         else:
#             date1 = data_start1.strftime("%Y%m%d%H%M")
#             open1 = ws.rows[row+4][1].value
#             high = ws.rows[row+4][2].value
#             low = ws.rows[row+4][3].value
#             last_price = ws.rows[row+4][4].value
#             number_ticks = ws.rows[row+4][5].value
#             volume = ws.rows[row+4][6].value
#             value = ws.rows[row+4][7].value
#             yd = date1[:8]
#             time = date1[8:]
#             H = name+date1
#             with open('/Users/wode/Desktop/test.csv', 'a')as f:
#                 writer = csv.writer(f)
#                 writer.writerow([H, name, yd, time, open1, high, low, last_price, number_ticks, volume, value])







# pre = datetime.datetime(2015, 12, 15, 9, 30)
# for row in range(row_num):
#
#     date = ws.rows[row][0].value
#     if date < datetime.datetime(2015, 12, 15, 9, 30):
#         continue
#     elif date == datetime.datetime(2015, 12, 15, 9, 30):
#         date1 == pre
#     date1 = ws.rows[row][0].strftime("%Y%m%d%H%M")
#     open = ws.rows[row][1].value
#     high = ws.rows[row][2].value
#     low = ws.rows[row][3].value
#     last_price = ws.rows[row][4].value
#     number_ticks = ws.rows[row][5].value
#     volume = ws.rows[row][6].value
#     value = ws.rows[row][7].value
#     yd = date1[:8]
#     time = date1[8:]
#     H = name+date1





